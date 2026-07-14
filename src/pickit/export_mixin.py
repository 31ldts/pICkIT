"""Exporting interaction data: DataFrame conversion and Excel output.

Extracted verbatim from the original monolithic ``analyze_interactions.py``
(Fase 2, paso 5 del plan de modularización). No logic changes were made
during this extraction — only the physical location of the code changed.

``_export_bar_data_to_excel`` and ``_export_pie_data_to_excel`` (the two
helpers that write bar/pie chart *data* to `.xlsx`, as opposed to the
`.png` rendering itself) live here per Anexo A, even though their only
callers (`bar_chart`, `pie_chart`) live in `plot_mixin.py` — cross-mixin
calls via `self` are expected at this stage of the modularization.
"""

import copy
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .exceptions import InvalidFileExtensionException
from .models import InteractionData


class ExportMixin:
    """DataFrame/Excel export operations for interaction data.

    Depends on ``ValidationMixin`` (``self._check_variable_types``,
    ``self._verify_dimensions``) and on ``self.saving_directory``.
    """

    def get_dataframe(self, interaction_data: InteractionData) -> pd.DataFrame:
        """
        Converts the interaction matrix into a pandas DataFrame.

        Args:
            interaction_data (InteractionData): The object containing the interaction matrix.

        Returns:
            pd.DataFrame: A DataFrame representation of the interaction matrix.
        """
        # Validate the type of interaction_data
        self._check_variable_types(
            variables=[interaction_data], expected_types=[InteractionData], variable_names=["interaction_data"]
        )

        # Convert the matrix to a DataFrame and set the index and columns
        df = pd.DataFrame(interaction_data.matrix[1:], columns=interaction_data.matrix[0])
        df.set_index(df.columns[0], inplace=True)

        return df

    def get_interactions(self, interaction_data: InteractionData) -> dict[int, str]:
        """
        Retrieves the interaction labels from the InteractionData object.

        Args:
            interaction_data (InteractionData): The object containing the interaction matrix.

        Returns:
            list[str]: A list of interaction labels.
        """
        # Validate the type of interaction_data
        self._check_variable_types(
            variables=[interaction_data], expected_types=[InteractionData], variable_names=["interaction_data"]
        )
        return {i + 1: interaction for i, interaction in enumerate(interaction_data.interactions)}

    def _export_bar_data_to_excel(self, filename, indices, data, interaction_labels):
        """
        Export bar-chart data to an Excel file.

        Args:
            filename (str): Output .xlsx filename.
            indices (list[str]): Labels of each row (residue or PDB ID).
            data (list[list[int]]): Matrix of counts (elements × interaction types).
            interaction_labels (list[str]): Names of each interaction type.
        """

        if not data:
            raise ValueError("No data to export.")
        n_cols = len(interaction_labels)
        for i, row in enumerate(data):
            if len(row) != n_cols:
                raise ValueError(
                    f"Row {i} has {len(row)} columns, expected {n_cols} (interaction_labels={interaction_labels})"
                )

        # Prepare DataFrame
        df_dict = {"Element": indices}

        for i, label in enumerate(interaction_labels):
            df_dict[label] = [row[i] for row in data]

        df = pd.DataFrame(df_dict)

        # Save to Excel
        df.to_excel(filename, index=False)
        print(f"Data successfully saved to {filename}")

    def _export_pie_data_to_excel(self, filename, labels, counts):
        """
        Export pìe-chart data to an Excel file.
        """
        total = sum(counts)
        df = pd.DataFrame(
            {"Interaction": labels, "Count": counts, "Percent": [(c / total * 100) if total else 0 for c in counts]}
        )
        df.to_excel(filename, index=False)
        print(f"Data successfully saved to {filename}")

    def save_interaction_data(self, interaction_data: InteractionData, filename: str) -> None:
        """
        Saves the interaction data to an Excel file in the specified directory.

        This method exports the interaction matrix and additional attributes to an Excel file
        with two separate sheets:
            - **Matrix**: Contains the interaction matrix.
            - **Attributes**: Stores metadata such as interaction types, colors, ligand status,
            processing mode, protein consideration, and subunit status.

        Args:
            interaction_data (InteractionData): The interaction data to be saved.
            filename (str): The name of the output file (must end with '.xlsx').

        Returns:
            None

        Raises:
            ValueError: If the matrix dimensions are not valid.
            InvalidFileExtensionException: If the filename does not end with '.xlsx'.
            TypeMismatchException: If a variable type doesn't match the expected one.
        """

        # Validate variable types
        self._check_variable_types(
            variables=[interaction_data, filename],
            expected_types=[InteractionData, str],
            variable_names=["interaction_data", "filename"],
        )

        # Ensure the filename ends with .xlsx
        if not filename.endswith(".xlsx"):
            raise InvalidFileExtensionException(filename)

        # Validate matrix dimensions
        self._verify_dimensions(matrix=interaction_data.matrix)

        # Construct the file path
        file_path = os.path.join(self.saving_directory, filename)

        # Prepare data for the first sheet (Matrix)
        if not (interaction_data.protein and interaction_data.ligand) and interaction_data.subunit:
            matrix = copy.deepcopy(interaction_data.matrix)
            for i in range(1, len(matrix)):
                for j in range(1, len(matrix[i])):
                    if matrix[i][j] != "":
                        matrix[i][j] = matrix[i][j].replace(" ||", "")
            matrix_df = pd.DataFrame(matrix)
        else:
            matrix_df = pd.DataFrame(interaction_data.matrix)

        # Prepare data for the second sheet (Attributes)
        # Create id, interacciones, colores columns for lists
        empty_strings = [""] * (
            len(interaction_data.interactions) - 1
        )  # Donde 'n' es el número de cadenas vacías que quieres.
        subunits_str = ", ".join(sorted(interaction_data.subunits_set))

        interactions_data = pd.DataFrame(
            {
                "id": range(1, len(interaction_data.interactions) + 1),
                "interactions": interaction_data.interactions,
                "colors": interaction_data.colors,
                "ligand": ["True" if interaction_data.ligand else "False"] + empty_strings,
                "mode": [interaction_data.mode] + empty_strings,
                "protein": ["True" if interaction_data.protein else "False"] + empty_strings,
                "subunit": ["True" if interaction_data.subunit else "False" + f" ({subunits_str})"] + empty_strings,
            }
        )

        # Save to an Excel file with two sheets
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # First sheet: Matrix
            matrix_df.to_excel(writer, sheet_name="Matrix", index=False, header=False)

            # Second sheet: Interactions & Attributes
            # Write interactions and colors
            interactions_data.to_excel(writer, sheet_name="Attributes", index=False, startrow=0)

        # Apply background colors to the row "colors"
        wb = load_workbook(file_path)
        ws = wb["Attributes"]

        col_index = 3  # Column "colors" (1-based index in Excel)

        for row_idx, color_hex in enumerate(interaction_data.colors, start=2):  # Starts on row 2 (skipping the header)
            fill = PatternFill(
                start_color=color_hex.replace("#", ""), end_color=color_hex.replace("#", ""), fill_type="solid"
            )
            ws.cell(row=row_idx, column=col_index).fill = fill

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2  # Añadir un poco de espacio extra

        wb.save(file_path)

        print(f"Interaction data successfully saved to {file_path}")
